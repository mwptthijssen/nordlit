"""
NordLit – Multi-database Nordic Literature Search
app.py  ·  FastAPI backend
"""
from __future__ import annotations

import csv
import html
import io
import json
import os
import re
from datetime import datetime
from typing import Any

from fastapi import FastAPI, Query, Request
from fastapi.responses import HTMLResponse, PlainTextResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from opensearchpy import OpenSearch

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

OPENSEARCH_URL = os.environ.get("OPENSEARCH_URL", "http://localhost:9200")
OPENSEARCH_USER = os.environ.get("OPENSEARCH_USER", "")
OPENSEARCH_PASS = os.environ.get("OPENSEARCH_PASS", "")
INDEX = os.environ.get("NORDLIT_INDEX", "nordlit")

EXCLUDED_SOURCES: set[str] = {"libris", "Libris", "libris_smart", "Libris_smart"}

# ---------------------------------------------------------------------------
# Source grouping and labels
# ---------------------------------------------------------------------------

SOURCE_LABELS: dict[str, str] = {
    "nva": "NVA (Norway)",
    "diva": "DiVA (Sweden)",
    "finna": "Finna (Finland)",
    "forskningsportal": "Forskningsportal (Denmark)",
    "helda": "Helda (Finland)",
    "journal_fi": "Journal.fi (Finland)",
    "opinvisindi": "Opin Visindi (Iceland)",
    "skemman": "Skemman (Iceland)",
    "theseus": "Theseus (Finland)",
    "tidsskrift_dk": "Tidsskrift.dk (Denmark)",
    "ojs_hi": "OJS (Iceland)",
    "openalex": "OpenAlex (Global)",
    "aalborg_open_journals": "Aalborg University Open Journals (Denmark)",
    "abo_journals": "Åbo Akademi OJS (Finland)",
    "aicvs": "Äldre I Centrum Vetenskapligt Supplement (Sweden)",
    "annals_of_representation_theory": "Annals of Representation Theory (Norway)",
    "arctic_portal_library": "Arctic Portal Library (Iceland)",
    "arctic_review": "Arctic Review on Law and Politics (Norway)",
    "arkiv_journal": "Arkiv. Tidskrift för samhällsanalys (Sweden)",
    "au_open_books": "AU Library E-books (Denmark)",
    "barnboken": "Barnboken (Sweden)",
    "bells": "BeLLS (Norway)",
    "bibsys": "BIBSYS Open Journals (Norway)",
    "boap_uib": "Bergen Open Access Publishing (Norway)",
    "cbs_open_journals": "CBS Open Journals (Denmark)",
    "drcmr_eprints": "DRCMR Eprints (Denmark)",
    "editori_helsinki": "Editori (Finland)",
    "ejedp": "European Journal of Economic Dynamics and Policy (Denmark)",
    "estetika": "Estetika (Finland)",
    "franorfon": "Revue nordique des études francophones (Sweden)",
    "geus_bulletin": "GEUS Bulletin (Denmark)",
    "gu_ojs": "University of Gothenburg OJS (Sweden)",
    "hprints": "hprints.org (Denmark)",
    "hup": "Helsinki University Press (Finland)",
    "iberoamericana": "Iberoamericana (Sweden)",
    "jdsr": "Journal of Digital Social Research (Sweden)",
    "jicc": "Journal of Intercultural Communication (Sweden)",
    "jisib": "Journal of Intelligence Studies in Business (Sweden)",
    "karib": "Karib (Sweden)",
    "landspitali_archive": "Landspitali Research Archive (Iceland)",
    "linkoping_ep": "Linköping University Electronic Press (Sweden)",
    "liu_electronic_press": "LiU Electronic Press (Sweden)",
    "lnuopen": "LnuOpen (Sweden)",
    "malmo_ojs": "OJS @ Malmö University (Sweden)",
    "maritime_commons": "Maritime Commons (Sweden)",
    "njas": "Nordic Journal of African Studies (Finland)",
    "njls": "Nordic Journal of Legal Studies (Finland)",
    "njmr": "Nordic Journal of Migration Research (Finland)",
    "nordic_academic_press": "Nordic Academic Press (Sweden)",
    "noril": "Nordic Journal of Information Literacy (Norway)",
    "novus_ebooks": "Novus E-bøker (Norway)",
    "novus_ojs": "Novus Online Tidsskrifter (Norway)",
    "ntnu_ojs": "NTNU Open Access Journals (Norway)",
    "ojlu": "OJLU – Lund University (Sweden)",
    "open_books_lund": "Open Books at Lund University (Sweden)",
    "orgprints": "Organic Eprints (Denmark)",
    "orkana": "Orkana Forlag (Norway)",
    "oslomet_journals": "OsloMet Open Access Journals (Norway)",
    "polar_research": "Polar Research (Norway)",
    "redescriptions": "Redescriptions (Finland)",
    "rural_landscapes": "Rural Landscapes (Sweden)",
    "septentrio": "Septentrio (Norway)",
    "silva_fennica_ojs": "Silva Fennica OJS (Finland)",
    "sjdr": "SJDR (Sweden)",
    "sjms": "Scandinavian Journal of Military Studies (Norway)",
    "sjwop": "SJWOP (Sweden)",
    "skriftserien_oslomet": "HiOA Skriftserien (Norway)",
    "socialmedicinsk_tidskrift": "Socialmedicinsk tidskrift (Sweden)",
    "stockholm_university_press": "Stockholm University Press (Sweden)",
    "textos_en_proceso": "Textos en Proceso (Sweden)",
    "tup": "Tampere University Press (Finland)",
    "uia_journals": "UiA Journal System (Norway)",
    "uio_fritt": "UiO FRITT (Norway)",
    "umea_journals": "Umeå University Hosted Journals (Sweden)",
    "universitetsforlaget": "Universitetsforlaget (Norway)",
    "vbri_press": "VBRI Press (Sweden)",
    "vtt_portal": "VTT Research Information Portal (Finland)",
}

PUBTYPE_LABELS: dict[str, str] = {
    "journal_article": "Journal Article",
    "review_article": "Review Article",
    "editorial": "Editorial",
    "letter_or_comment": "Letter or Comment",
    "note_or_short_communication": "Note / Short Communication",
    "registered_report": "Registered Report",
    "conference_paper": "Conference Paper",
    "conference_abstract": "Conference Abstract",
    "conference_poster": "Conference Poster",
    "conference_proceeding": "Conference Proceeding",
    "book": "Book",
    "edited_book": "Edited Book",
    "book_chapter": "Book Chapter",
    "report": "Report",
    "working_paper": "Working Paper",
    "policy_brief": "Policy Brief",
    "doctoral_dissertation": "Doctoral Dissertation",
    "licentiate_thesis": "Licentiate Thesis",
    "masters_thesis": "Master's Thesis",
    "bachelors_thesis": "Bachelor's Thesis",
    "student_thesis_unspecified": "Student Thesis (unspecified level)",
    "preprint": "Preprint",
    "dataset": "Dataset",
    "software": "Software",
    "patent": "Patent",
    "encyclopedia_entry": "Encyclopedia Entry",
    "reference_entry": "Reference Entry",
    "magazine_article": "Magazine Article",
    "newspaper_article": "Newspaper Article",
    "popular_science_article": "Popular Science Article",
    "professional_article": "Professional Article",
    "media_contribution": "Media Contribution",
    "artistic_output": "Artistic Output",
    "presentation_or_lecture": "Presentation or Lecture",
    "other": "Other",
    "unknown": "Unknown",
    "missing": "Missing",
}


def _prettify_source_key(key: str) -> str:
    key = (key or "").strip().replace("_", " ")
    return " ".join(part.capitalize() for part in key.split())


def _source_group(key: str) -> str:
    key = key or ""
    key_lc = key.lower()
    if key_lc.startswith("swepub"):
        return "SwePub (Sweden)"
    if key_lc.startswith("fi_institutions"):
        return "Finnish institutions (Finland)"
    if key_lc in {k.lower() for k in DEFAULT_SOURCE_GROUPS.get("Journal Platforms", [])}:
        return "Journal Platforms"
    return SOURCE_LABELS.get(key, SOURCE_LABELS.get(key_lc, _prettify_source_key(key)))


DEFAULT_SOURCE_GROUPS: dict[str, list[str]] = {
    "NVA (Norway)": ["nva"],
    "Polar Research (Norway)": ["polar_research"],
    "DiVA (Sweden)": ["diva"],
    "SwePub (Sweden)": ["swepub"],
    "Maritime Commons (Sweden)": ["maritime_commons"],
    "LiU Electronic Press (Sweden)": ["liu_electronic_press"],
    "Open Books at Lund University (Sweden)": ["open_books_lund"],
    "Stockholm University Press (Sweden)": ["stockholm_university_press"],
    "Forskningsportal (Denmark)": ["forskningsportal"],
    "Organic Eprints (Denmark)": ["orgprints"],
    "AU Library E-books (Denmark)": ["au_open_books"],
    "hprints.org (Denmark)": ["hprints"],
    "Finna (Finland)": ["finna"],
    "Finnish institutions (Finland)": ["fi_institutions"],
    "Theseus (Finland)": ["theseus"],
    "Helda (Finland)": ["helda"],
    "Skemman (Iceland)": ["skemman"],
    "Landspitali Research Archive (Iceland)": ["landspitali_archive"],
    "Opin Visindi (Iceland)": ["opinvisindi"],
    "Arctic Portal Library (Iceland)": ["arctic_portal_library"],
    "OpenAlex (Global)": ["openalex"],
    "Journal Platforms": [
        "abo_journals",
        "aicvs",
        "uio_fritt",
        "septentrio",
        "uia_journals",
        "umea_journals",
        "textos_en_proceso",
        "socialmedicinsk_tidskrift",
        "sjwop",
        "sjms",
        "sjdr",
        "rural_landscapes",
        "redescriptions",
        "oslomet_journals",
        "ojlu",
        "gu_ojs",
        "malmo_ojs",
        "ntnu_ojs",
        "novus_ojs",
        "njmr",
        "noril",
        "lnuopen",
        "linkoping_ep",
        "jdsr",
        "skriftserien_oslomet",
        "geus_bulletin",
        "silva_fennica_ojs",
        "estetika",
        "editori_helsinki",
        "cbs_open_journals",
        "boap_uib",
        "bells",
        "barnboken",
        "arkiv_journal",
        "arctic_review",
        "annals_of_representation_theory",
        "ojs_hi",
        "journal_fi",
        "tidsskrift_dk",
        "njls",
        "aalborg_open_journals",
    ],
}

GROUP_ORDER: list[str] = [
    "NVA (Norway)",
    "Polar Research (Norway)",
    "DiVA (Sweden)",
    "SwePub (Sweden)",
    "Maritime Commons (Sweden)",
    "LiU Electronic Press (Sweden)",
    "Open Books at Lund University (Sweden)",
    "Stockholm University Press (Sweden)",
    "Forskningsportal (Denmark)",
    "Organic Eprints (Denmark)",
    "AU Library E-books (Denmark)",
    "hprints.org (Denmark)",
    "Finna (Finland)",
    "Finnish institutions (Finland)",
    "Theseus (Finland)",
    "Helda (Finland)",
    "Skemman (Iceland)",
    "Landspitali Research Archive (Iceland)",
    "Opin Visindi (Iceland)",
    "Arctic Portal Library (Iceland)",
    "OpenAlex (Global)",
    "Journal Platforms",
]
# Country order for Journal Platforms sub-item sorting
JOURNAL_COUNTRY_ORDER: dict[str, int] = {
    "Norway":  0,
    "Sweden":  1,
    "Denmark": 2,
    "Finland": 3,
    "Iceland": 4,
}

# Map individual journal source keys to their country
JOURNAL_SOURCE_COUNTRY: dict[str, str] = {
    # Norway
    "uio_fritt": "Norway",
    "septentrio": "Norway",
    "uia_journals": "Norway",
    "oslomet_journals": "Norway",
    "ntnu_ojs": "Norway",
    "novus_ojs": "Norway",
    "noril": "Norway",
    "skriftserien_oslomet": "Norway",
    "boap_uib": "Norway",
    "bells": "Norway",
    "arctic_review": "Norway",
    "annals_of_representation_theory": "Norway",
    "sjms": "Norway",
    # Sweden
    "aicvs": "Sweden",
    "umea_journals": "Sweden",
    "textos_en_proceso": "Sweden",
    "socialmedicinsk_tidskrift": "Sweden",
    "sjwop": "Sweden",
    "sjdr": "Sweden",
    "rural_landscapes": "Sweden",
    "ojlu": "Sweden",
    "gu_ojs": "Sweden",
    "malmo_ojs": "Sweden",
    "lnuopen": "Sweden",
    "linkoping_ep": "Sweden",
    "jdsr": "Sweden",
    "barnboken": "Sweden",
    "arkiv_journal": "Sweden",
    # Denmark
    "geus_bulletin": "Denmark",
    "cbs_open_journals": "Denmark",
    "tidsskrift_dk": "Denmark",
    "aalborg_open_journals": "Denmark",
    # Finland
    "abo_journals": "Finland",
    "redescriptions": "Finland",
    "silva_fennica_ojs": "Finland",
    "estetika": "Finland",
    "editori_helsinki": "Finland",
    "njmr": "Finland",
    "njls": "Finland",
    "journal_fi": "Finland",
    # Iceland
    "ojs_hi": "Iceland",
}

GROUP_ORDER_INDEX: dict[str, int] = {label: idx for idx, label in enumerate(GROUP_ORDER)}

# ---------------------------------------------------------------------------
# HTML cleaning
# ---------------------------------------------------------------------------

_HTML_TAG_RE = re.compile(r"<[^>]+>")
_MULTI_SPACE_RE = re.compile(r"\s{2,}")


def clean_html(text: str) -> str:
    if not text:
        return ""
    text = html.unescape(str(text))
    text = _HTML_TAG_RE.sub(" ", text)
    text = _MULTI_SPACE_RE.sub(" ", text)
    return text.strip()


# ---------------------------------------------------------------------------
# OpenSearch client
# ---------------------------------------------------------------------------

def _make_client() -> OpenSearch:
    kwargs: dict[str, Any] = {"hosts": [OPENSEARCH_URL], "timeout": 30}
    if OPENSEARCH_USER and OPENSEARCH_PASS:
        kwargs["http_auth"] = (OPENSEARCH_USER, OPENSEARCH_PASS)
    if OPENSEARCH_URL.startswith("https"):
        kwargs["use_ssl"] = True
        kwargs["verify_certs"] = False
    return OpenSearch(**kwargs)


client = _make_client()

# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------

app = FastAPI(title="NordLit Search")
if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory=".")


# ---------------------------------------------------------------------------
# Query builders
# ---------------------------------------------------------------------------

OTHER_YEAR_MIN = 2033
OTHER_YEAR_MAX = 2099

def _build_year_filters(
    year_from: int | None,
    year_to: int | None,
    years_exact: list[str],
) -> list[dict]:
    filters: list[dict] = []
    year_clauses: list[dict] = []

    if year_from or year_to:
        rc: dict[str, int] = {}
        if year_from is not None:
            rc["gte"] = year_from
        if year_to is not None:
            rc["lte"] = year_to
        year_clauses.append({"range": {"year": rc}})

    for y in years_exact:
        if str(y) == "Other":
            # Records with year 2033-2099 OR with no year value at all
            year_clauses.append({"bool": {"should": [
                {"range": {"year": {"gte": OTHER_YEAR_MIN, "lte": OTHER_YEAR_MAX}}},
                {"bool": {"must_not": {"exists": {"field": "year"}}}},
            ], "minimum_should_match": 1}})
        else:
            try:
                year_clauses.append({"term": {"year": int(y)}})
            except (ValueError, TypeError):
                pass

    if year_clauses:
        if len(year_clauses) == 1:
            filters.append(year_clauses[0])
        else:
            filters.append({"bool": {"should": year_clauses, "minimum_should_match": 1}})

    return filters


def _text_clause(q: str) -> list[dict]:
    if not q.strip():
        return [{"match_all": {}}]
    return [{
        "query_string": {
            "query": q,
            "fields": ["title^3", "abstract"],
            "default_operator": "AND",
        }
    }]


def _unique_preserve(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        if not item or item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def _parse_source_pubtype_pairs(source_pubtypes: list[str]) -> dict[str, set[str]]:
    mapping: dict[str, set[str]] = {}
    for value in source_pubtypes or []:
        if not value or "::" not in value:
            continue
        source_key, pubtype_key = value.split("::", 1)
        if not source_key or not pubtype_key or source_key in EXCLUDED_SOURCES:
            continue
        mapping.setdefault(source_key, set()).add(pubtype_key)
    return mapping


def _build_query(
    q: str,
    sources: list[str],
    pub_types: list[str],
    source_pubtypes: list[str],
    year_from: int | None,
    year_to: int | None,
    years_exact: list[str],
) -> dict:
    must = _text_clause(q)
    filters: list[dict] = []

    selected_sources = _unique_preserve([
        source for source in (sources or [])
        if source and source not in EXCLUDED_SOURCES
    ])
    scoped_pubtypes = _parse_source_pubtype_pairs(source_pubtypes)
    effective_sources = _unique_preserve(selected_sources + list(scoped_pubtypes.keys()))

    if effective_sources:
        if scoped_pubtypes:
            scoped_sources = set(scoped_pubtypes.keys())
            should_clauses: list[dict] = []
            unrestricted_sources = [source for source in effective_sources if source not in scoped_sources]
            if unrestricted_sources:
                should_clauses.append({"terms": {"source": unrestricted_sources}})
            for source_key in _unique_preserve(list(scoped_pubtypes.keys())):
                should_clauses.append({
                    "bool": {
                        "filter": [
                            {"term": {"source": source_key}},
                            {"terms": {"publication_type_normalized": sorted(scoped_pubtypes[source_key])}},
                        ]
                    }
                })
            if should_clauses:
                filters.append({"bool": {"should": should_clauses, "minimum_should_match": 1}})
        else:
            filters.append({"terms": {"source": effective_sources}})

    legacy_pub_types = _unique_preserve(pub_types or [])
    if legacy_pub_types and not scoped_pubtypes:
        filters.append({"terms": {"publication_type_normalized": legacy_pub_types}})

    filters.extend(_build_year_filters(year_from, year_to, years_exact))

    query: dict = {"bool": {"must": must}}
    if filters:
        query["bool"]["filter"] = filters
    query["bool"]["must_not"] = [{"terms": {"source": sorted(EXCLUDED_SOURCES)}}]
    return query


def _build_facet_query(
    q: str,
    year_from: int | None,
    year_to: int | None,
    years_exact: list[str],
) -> dict:
    query: dict = {"bool": {"must": _text_clause(q)}}
    year_filters = _build_year_filters(year_from, year_to, years_exact)
    if year_filters:
        query["bool"]["filter"] = year_filters
    query["bool"]["must_not"] = [{"terms": {"source": sorted(EXCLUDED_SOURCES)}}]
    return query


def _search(q, sources, pub_types, source_pubtypes, year_from, year_to, years_exact, page, size) -> dict:
    hits_body = {
        "query": _build_query(q, sources, pub_types, source_pubtypes, year_from, year_to, years_exact),
        "sort": [{"year": {"order": "desc", "missing": "_last"}}, "_score"],
        "from": (page - 1) * size,
        "size": size,
        "track_total_hits": True,
    }
    facet_body = {
        "query": _build_facet_query(q, year_from, year_to, years_exact),
        "size": 0,
        "aggs": {
            "sources": {"terms": {"field": "source", "size": 100}},
            "source_pubtypes": {
                "terms": {"field": "source", "size": 100},
                "aggs": {"pubtypes": {"terms": {"field": "publication_type_normalized", "size": 100}}},
            },
            "years": {"terms": {"field": "year", "size": 200, "order": {"_key": "desc"}}},
            "years_missing": {"missing": {"field": "year"}},
        },
    }
    hits = client.search(index=INDEX, body=hits_body)
    facets = client.search(index=INDEX, body=facet_body)
    hits["aggregations"] = facets.get("aggregations", {})
    return hits


# ---------------------------------------------------------------------------
# Source-group aggregation helper
# ---------------------------------------------------------------------------

def _group_sources(source_buckets, source_pubtype_buckets):
    groups: dict[str, dict[str, Any]] = {}

    for label, default_keys in DEFAULT_SOURCE_GROUPS.items():
        groups[label] = {
            "group_label": label,
            "keys": [key for key in default_keys if key not in EXCLUDED_SOURCES],
            "count": 0,
            "pubtypes": {},
        }

    for b in source_buckets:
        key = b["key"]
        if key in EXCLUDED_SOURCES:
            continue
        label = _source_group(key)
        if label not in groups:
            groups[label] = {"group_label": label, "keys": [], "count": 0, "pubtypes": {}}
        if key not in groups[label]["keys"]:
            groups[label]["keys"].append(key)
        groups[label]["count"] += b["doc_count"]

    # Track per-source counts for Journal Platforms (source label as sub-item)
    journal_platform_sources: dict[str, dict] = {}

    for b in source_pubtype_buckets:
        key = b["key"]
        if key in EXCLUDED_SOURCES:
            continue
        label = _source_group(key)
        if label not in groups:
            groups[label] = {"group_label": label, "keys": [key], "count": 0, "pubtypes": {}}
        elif key not in groups[label]["keys"]:
            groups[label]["keys"].append(key)

        if label == "Journal Platforms":
            # For Journal Platforms: sub-items are individual journals, not pub types
            source_count = sum(pt["doc_count"] for pt in b.get("pubtypes", {}).get("buckets", []))
            source_label = SOURCE_LABELS.get(key, _prettify_source_key(key))
            if key not in journal_platform_sources:
                journal_platform_sources[key] = {
                    "key": key,
                    "label": source_label,
                    "count": 0,
                }
            journal_platform_sources[key]["count"] += source_count
        else:
            for pt in b.get("pubtypes", {}).get("buckets", []):
                pt_key = pt["key"]
                if pt_key not in groups[label]["pubtypes"]:
                    groups[label]["pubtypes"][pt_key] = {
                        "key": pt_key,
                        "label": PUBTYPE_LABELS.get(pt_key, pt_key),
                        "count": 0,
                    }
                groups[label]["pubtypes"][pt_key]["count"] += pt["doc_count"]

    # Inject journal sources as the sub-items for Journal Platforms
    if "Journal Platforms" in groups and journal_platform_sources:
        groups["Journal Platforms"]["pubtypes"] = journal_platform_sources
        groups["Journal Platforms"]["show_sources"] = True

    result = []
    for g in sorted(groups.values(), key=lambda x: (GROUP_ORDER_INDEX.get(x["group_label"], 999), x["group_label"].lower())):
        if not g.get("show_sources"):
            g["pubtypes"] = sorted(
                g["pubtypes"].values(),
                key=lambda x: (-x["count"], x["label"].lower()),
            )
        else:
            g["pubtypes"] = sorted(
                g["pubtypes"].values(),
                key=lambda x: (
                    JOURNAL_COUNTRY_ORDER.get(JOURNAL_SOURCE_COUNTRY.get(x["key"], ""), 99),
                    -x["count"],
                ),
            )
        result.append(g)
    return result


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

def _all_hits(query: dict):
    body = {
        "query": query,
        "sort": [{"year": {"order": "desc", "missing": "_last"}}],
        "size": 1000,
    }
    resp = client.search(index=INDEX, body=body, scroll="5m")
    scroll_id = resp.get("_scroll_id")
    hits = resp["hits"]["hits"]
    while hits:
        yield from hits
        resp = client.scroll(scroll_id=scroll_id, scroll="5m")
        scroll_id = resp.get("_scroll_id")
        hits = resp["hits"]["hits"]
    if scroll_id:
        try:
            client.clear_scroll(scroll_id=scroll_id)
        except Exception:
            pass


def _fmt_src(h: dict) -> dict:
    s = h["_source"]
    return {
        "title": clean_html(s.get("title", "")),
        "authors": clean_html(s.get("authors", "")),
        "year": s.get("year", ""),
        "abstract": clean_html(s.get("abstract", "")),
        "source": _source_group(s.get("source", "")),
        "pub_type": PUBTYPE_LABELS.get(
            s.get("publication_type_normalized", ""),
            s.get("publication_type_normalized", ""),
        ),
        "pub_type_raw": s.get("publication_type", ""),
        "source_url": s.get("source_url", ""),
    }


def _count_records(query: dict, ids: list[str] | None = None) -> int:
    if ids:
        return len(ids)
    try:
        return int(client.count(index=INDEX, body={"query": query}).get("count", 0))
    except Exception:
        count = 0
        for _ in _all_hits(query):
            count += 1
        return count


def _export_base_name(record_count: int) -> str:
    today = datetime.now().strftime("%d_%m_%Y")
    return f"Nordlit_{today}_{record_count}_records"


def _content_disposition(filename: str) -> str:
    return f'attachment; filename="{filename}"'


EXPORT_FIELDS = ["title", "authors", "year", "abstract", "source", "pub_type", "pub_type_raw", "source_url"]


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html")


@app.get("/manual")
async def manual():
    """Serve manual.pdf for download/viewing in a new tab."""
    path = "manual.pdf"
    if not os.path.isfile(path):
        return PlainTextResponse("Manual not found.", status_code=404)
    def iter_file():
        with open(path, "rb") as f:
            while chunk := f.read(65536):
                yield chunk
    return StreamingResponse(
        iter_file(),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="NordLit_Manual.pdf"'},
    )


@app.get("/api/search")
async def search(
    q: str = Query(default=""),
    sources: list[str] = Query(default=[]),
    pub_types: list[str] = Query(default=[]),
    pub_type_pairs: list[str] = Query(default=[]),
    year_from: int | None = Query(default=None),
    year_to: int | None = Query(default=None),
    years_exact: list[str] = Query(default=[]),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=25),
):
    raw = _search(q, sources, pub_types, pub_type_pairs, year_from, year_to, years_exact, page, size)
    total = raw["hits"]["total"]["value"]
    hits = raw["hits"]["hits"]

    results = []
    for h in hits:
        s = h["_source"]
        results.append({
            "id": h["_id"],
            "title": clean_html(s.get("title", "")),
            "authors": clean_html(s.get("authors", "")),
            "year": s.get("year"),
            "abstract": clean_html(s.get("abstract", "")),
            "source": s.get("source", ""),
            "source_label": _source_group(s.get("source", "")),
            "pub_type_raw": s.get("publication_type", ""),
            "pub_type_normalized": s.get("publication_type_normalized", ""),
            "pub_type_label": PUBTYPE_LABELS.get(
                s.get("publication_type_normalized", ""),
                s.get("publication_type_normalized", ""),
            ),
            "source_url": s.get("source_url", ""),
        })

    aggs = raw.get("aggregations", {})
    grouped = _group_sources(
        aggs.get("sources", {}).get("buckets", []),
        aggs.get("source_pubtypes", {}).get("buckets", []),
    )
    other_count = sum(
        b["doc_count"]
        for b in aggs.get("years", {}).get("buckets", [])
        if b.get("key") and OTHER_YEAR_MIN <= int(b["key"]) <= OTHER_YEAR_MAX
    )
    other_count += aggs.get("years_missing", {}).get("doc_count", 0)
    year_buckets = []
    if other_count > 0:
        year_buckets.append({"year": "Other", "count": other_count})
    year_buckets += [
        {"year": b["key"], "count": b["doc_count"]}
        for b in aggs.get("years", {}).get("buckets", [])
        if b.get("key") and not (OTHER_YEAR_MIN <= int(b["key"]) <= OTHER_YEAR_MAX)
    ]

    return {
        "total": total,
        "page": page,
        "size": size,
        "results": results,
        "facets": {
            "source_groups": grouped,
            "years": year_buckets,
        },
    }


@app.get("/api/export")
async def export(
    fmt: str = Query(default="csv"),
    q: str = Query(default=""),
    sources: list[str] = Query(default=[]),
    pub_types: list[str] = Query(default=[]),
    pub_type_pairs: list[str] = Query(default=[]),
    year_from: int | None = Query(default=None),
    year_to: int | None = Query(default=None),
    years_exact: list[str] = Query(default=[]),
    ids: list[str] = Query(default=[]),
):
    fmt = (fmt or "csv").lower()

    if ids:
        query: dict = {"ids": {"values": ids}}
    else:
        query = _build_query(q, sources, pub_types, pub_type_pairs, year_from, year_to, years_exact)

    record_count = _count_records(query, ids if ids else None)
    base_name = _export_base_name(record_count)

    def hits():
        yield from _all_hits(query)

    if fmt == "csv":
        def gen_csv():
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS)
            w.writeheader()
            yield buf.getvalue()
            for h in hits():
                buf = io.StringIO()
                w = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS)
                w.writerow(_fmt_src(h))
                yield buf.getvalue()

        filename = f"{base_name}.csv"
        return StreamingResponse(
            gen_csv(),
            media_type="text/csv",
            headers={"Content-Disposition": _content_disposition(filename)},
        )

    if fmt == "jsonl":
        def gen_jsonl():
            for h in hits():
                yield json.dumps(h["_source"], ensure_ascii=False) + "\n"

        filename = f"{base_name}.jsonl"
        return StreamingResponse(
            gen_jsonl(),
            media_type="application/x-ndjson",
            headers={"Content-Disposition": _content_disposition(filename)},
        )

    if fmt == "bibtex":
        def _bib(rec: dict, idx: int) -> str:
            title = rec["title"].replace("{", "").replace("}", "")
            authors = rec["authors"].replace(";", " and ")
            year = str(rec["year"]) if rec["year"] else ""
            key = re.sub(r"\W+", "", authors.split(",")[0] if authors else "anon") + year + str(idx)
            pt = rec.get("pub_type", "")
            btype = (
                "phdthesis" if ("Thesis" in pt or "Dissertation" in pt)
                else "book" if "Book" in pt
                else "inproceedings" if "Conference" in pt
                else "techreport" if "Report" in pt
                else "article"
            )
            ls = [f"@{btype}{{{key},", f"  title  = {{{title}}},"]
            if authors:
                ls.append(f"  author = {{{authors}}},")
            if year:
                ls.append(f"  year   = {{{year}}},")
            if rec.get("source_url"):
                ls.append(f"  url    = {{{rec['source_url']}}},")
            if rec.get("abstract"):
                ls.append(f"  abstract = {{{rec['abstract'].replace('{', '').replace('}', '')[:800]}}},")
            ls.append("}")
            return "\n".join(ls)

        def gen_bib():
            for i, h in enumerate(hits()):
                yield _bib(_fmt_src(h), i) + "\n\n"

        filename = f"{base_name}.bib"
        return StreamingResponse(
            gen_bib(),
            media_type="text/plain",
            headers={"Content-Disposition": _content_disposition(filename)},
        )

    if fmt == "ris":
        def _ris(rec: dict) -> str:
            pt = rec.get("pub_type", "")
            ty = (
                "THES" if ("Thesis" in pt or "Dissertation" in pt)
                else "CHAP" if "Book Chapter" in pt
                else "BOOK" if "Book" in pt
                else "CONF" if "Conference" in pt
                else "RPRT" if "Report" in pt
                else "JOUR"
            )
            ls = [f"TY  - {ty}", f"TI  - {rec['title']}"]
            for a in rec["authors"].split(";"):
                a = a.strip()
                if a:
                    ls.append(f"AU  - {a}")
            if rec["year"]:
                ls.append(f"PY  - {rec['year']}")
            if rec.get("abstract"):
                ls.append(f"AB  - {rec['abstract'][:800]}")
            if rec.get("source_url"):
                ls.append(f"UR  - {rec['source_url']}")
            ls.append(f"DP  - {rec.get('source', '')}")
            ls.append("ER  - ")
            return "\n".join(ls) + "\n\n"

        def gen_ris():
            for h in hits():
                yield _ris(_fmt_src(h))

        filename = f"{base_name}.ris"
        return StreamingResponse(
            gen_ris(),
            media_type="application/x-research-info-systems",
            headers={"Content-Disposition": _content_disposition(filename)},
        )

    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        rows_buf = [_fmt_src(h) for h in hits()]
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NordLit Export"
        headers = ["Title", "Authors", "Year", "Abstract", "Source", "Publication Type", "Raw Type", "URL"]
        ws.append(headers)
        hfill = PatternFill("solid", fgColor="1E3A5F")
        hfont = Font(bold=True, color="FFFFFF", name="Calibri")
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
        for r in rows_buf:
            ws.append([r["title"], r["authors"], r["year"], r["abstract"], r["source"], r["pub_type"], r["pub_type_raw"], r["source_url"]])
        for col in ws.columns:
            mx = max(len(str(c.value or "")) for c in col[:100])
            ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, 12), 60)
        wb.save(buf)
        buf.seek(0)
        filename = f"{base_name}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": _content_disposition(filename)},
        )

    return PlainTextResponse("Unsupported format", status_code=400)
